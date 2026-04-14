"""
excel_writer.py — Writes keyword research results to a formatted Excel file.
"""
import logging
import os
import re
from datetime import date
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_DIR

logger = logging.getLogger(__name__)


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Return a valid Excel sheet name (max 31 chars, no special chars)."""
    # Remove characters Excel forbids: / \ ? * [ ] :
    name = re.sub(r"[/\\?*\[\]:]", " ", name).strip()
    return name[:max_len] if len(name) > max_len else name


def _domain_slug(url: str) -> str:
    """Convert a URL to a safe filename slug, e.g. example_com."""
    netloc = urlparse(url).netloc
    return re.sub(r"[^a-z0-9]", "_", netloc.lower()).strip("_")


def _auto_width(ws) -> None:
    """Set column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def write_single_page(
    url: str,
    keyword_volumes: dict[str, dict],
) -> str:
    """
    Write a single-sheet Excel file for --url mode.

    Args:
        url: the source URL
        keyword_volumes: {keyword: {"volume": int, "competition": str}}

    Returns:
        Path to the generated .xlsx file.
    """
    domain_slug = _domain_slug(url)
    today = date.today().isoformat()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, f"{domain_slug}_{today}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active

    # Derive sheet name from URL path
    path = urlparse(url).path.rstrip("/")
    sheet_name = path.split("/")[-1] if "/" in path else domain_slug
    sheet_name = sheet_name.replace("-", " ").replace("_", " ").title()
    ws.title = _safe_sheet_name(sheet_name or domain_slug)

    _write_sheet(ws, url, keyword_volumes)

    wb.save(filepath)
    logger.info(f"Excel saved: {filepath}")
    return filepath


def write_multi_category(
    site_url: str,
    categories: dict[str, list[dict]],
) -> str:
    """
    Write a multi-sheet Excel file for --site mode.

    Args:
        site_url: the root site URL
        categories: {category_name: [{"keyword", "volume", "competition"}, ...]}

    Returns:
        Path to the generated .xlsx file.
    """
    domain_slug = _domain_slug(site_url)
    today = date.today().isoformat()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, f"{domain_slug}_{today}.xlsx")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for category_name, kw_list in categories.items():
        sheet_name = _safe_sheet_name(category_name)
        ws = wb.create_sheet(title=sheet_name)

        # Build keyword_volumes dict from list
        kv = {
            entry["keyword"]: {
                "volume": entry["volume"],
                "competition": entry.get("competition", "UNKNOWN"),
            }
            for entry in kw_list
        }
        _write_sheet(ws, site_url, kv, category_name=category_name)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Results")

    wb.save(filepath)
    logger.info(f"Excel saved: {filepath}")
    return filepath


def _write_sheet(
    ws,
    source_url: str,
    keyword_volumes: dict[str, dict],
    category_name: str = "",
) -> None:
    """Write data into a worksheet."""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    url_font = Font(color="0563C1", underline="single")

    # Row 1: source URL
    label = f"Source: {source_url}"
    if category_name:
        label = f"Category: {category_name} | Source: {source_url}"
    ws.append([label])
    ws["A1"].font = url_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    # Row 2: headers
    ws.append(["Keyword", "Volume", "Competition"])
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Rows 3+: keywords sorted by volume descending
    sorted_kws = sorted(
        keyword_volumes.items(),
        key=lambda x: -x[1].get("volume", 0),
    )
    for kw, data in sorted_kws:
        volume = data.get("volume", 0)
        competition = data.get("competition", "UNKNOWN")
        ws.append([kw, volume, competition])

    # Number format for volume column
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "#,##0"

    # Auto-width columns
    _auto_width(ws)
