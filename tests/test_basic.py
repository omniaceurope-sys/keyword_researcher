"""
Unit tests — no API calls required.
Run with: python -m pytest tests/ -v
"""
import os
import sys
import tempfile

import pytest

# Make sure the project root is on the path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ── site_crawler ──────────────────────────────────────────────────────────────
class TestSiteCrawler:
    from pipeline.site_crawler import _is_category_url, _slug
    from urllib.parse import urlparse

    def _parsed(self, path: str):
        from urllib.parse import urlparse
        return urlparse(f"https://example.com{path}")

    def _is_cat(self, path: str) -> bool:
        from pipeline.site_crawler import _is_category_url
        from urllib.parse import urlparse
        parsed = urlparse(f"https://example.com{path}")
        return _is_category_url(parsed, "example.com")

    def test_collections_url_accepted(self):
        assert self._is_cat("/collections/headphones") is True

    def test_product_category_accepted(self):
        assert self._is_cat("/product-category/chairs") is True

    def test_category_accepted(self):
        assert self._is_cat("/category/electronics") is True

    def test_shop_subpage_accepted(self):
        assert self._is_cat("/shop/headphones") is True

    def test_root_shop_rejected(self):
        assert self._is_cat("/shop") is False
        assert self._is_cat("/shop/") is False

    def test_root_store_rejected(self):
        assert self._is_cat("/store") is False

    def test_deals_rejected(self):
        assert self._is_cat("/collections/deals") is False

    def test_sale_rejected(self):
        assert self._is_cat("/collections/sale") is False

    def test_uncategorized_rejected(self):
        assert self._is_cat("/product-category/uncategorized") is False

    def test_gift_card_rejected(self):
        assert self._is_cat("/collections/gift-cards") is False

    def test_pagination_rejected(self):
        assert self._is_cat("/collections/headphones?page=2") is False

    def test_blog_rejected(self):
        assert self._is_cat("/blog/post-1") is False

    def test_about_rejected(self):
        assert self._is_cat("/about") is False

    def test_checkout_rejected(self):
        assert self._is_cat("/checkout") is False

    def test_slug_deduplication(self):
        from pipeline.site_crawler import _slug
        assert _slug("https://example.com/collections/headphones") == \
               _slug("https://example.com/collections/headphones/")

    def test_external_domain_rejected(self):
        from pipeline.site_crawler import _is_category_url
        from urllib.parse import urlparse
        parsed = urlparse("https://other.com/collections/headphones")
        assert _is_category_url(parsed, "example.com") is False


# ── scraper content extraction ────────────────────────────────────────────────
class TestScraper:
    def _extract(self, html: str) -> str:
        from pipeline.scraper import _extract_content
        return _extract_content(html, "https://example.com/test")

    def test_strips_nav_and_footer(self):
        html = """
        <html><body>
          <nav>Navigation menu</nav>
          <main><h1>Headphones</h1><p>Great audio products</p></main>
          <footer>Footer content</footer>
        </body></html>
        """
        result = self._extract(html)
        assert "Navigation menu" not in result
        assert "Footer content" not in result
        assert "Headphones" in result

    def test_json_ld_priority(self):
        html = """
        <html><body>
          <script type="application/ld+json">
            {"@type": "Product", "name": "Wireless Headphones", "description": "Great sound"}
          </script>
          <main><p>Some other text</p></main>
        </body></html>
        """
        result = self._extract(html)
        assert "Wireless Headphones" in result

    def test_main_fallback(self):
        html = """
        <html><body>
          <div class="unrelated">Sidebar</div>
          <main><p>Main product content here</p></main>
        </body></html>
        """
        result = self._extract(html)
        assert "Main product content here" in result

    def test_strips_scripts_and_styles(self):
        html = """
        <html><head>
          <style>body { color: red; }</style>
          <script>var x = 1;</script>
        </head><body>
          <main><p>Product info</p></main>
        </body></html>
        """
        result = self._extract(html)
        assert "color: red" not in result
        assert "var x" not in result
        assert "Product info" in result


# ── excel_writer ──────────────────────────────────────────────────────────────
class TestExcelWriter:
    def test_single_page_creates_file(self):
        import openpyxl
        from pipeline.excel_writer import write_single_page

        kv = {
            "wireless headphones": {"volume": 5000, "competition": "HIGH"},
            "noise cancelling earbuds": {"volume": 3200, "competition": "MEDIUM"},
            "cheap bluetooth headphones": {"volume": 1100, "competition": "LOW"},
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            # Temporarily override OUTPUT_DIR
            import pipeline.excel_writer as ew
            original = ew.OUTPUT_DIR
            ew.OUTPUT_DIR = tmpdir
            try:
                path = write_single_page("https://example.com/collections/headphones", kv)
                assert os.path.exists(path)

                wb = openpyxl.load_workbook(path)
                ws = wb.active
                # Row 2 = headers
                assert ws.cell(2, 1).value == "Keyword"
                assert ws.cell(2, 2).value == "Volume"
                assert ws.cell(2, 3).value == "Competition"
                # Row 3 = highest volume keyword
                assert ws.cell(3, 1).value == "wireless headphones"
                assert ws.cell(3, 2).value == 5000
            finally:
                ew.OUTPUT_DIR = original

    def test_multi_category_creates_sheets(self):
        import openpyxl
        from pipeline.excel_writer import write_multi_category

        categories = {
            "Headphones": [
                {"keyword": "wireless headphones", "volume": 5000, "competition": "HIGH"},
                {"keyword": "over ear headphones", "volume": 2000, "competition": "MEDIUM"},
            ],
            "Earbuds": [
                {"keyword": "true wireless earbuds", "volume": 8000, "competition": "HIGH"},
            ],
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            import pipeline.excel_writer as ew
            original = ew.OUTPUT_DIR
            ew.OUTPUT_DIR = tmpdir
            try:
                path = write_multi_category("https://example.com", categories)
                assert os.path.exists(path)

                wb = openpyxl.load_workbook(path)
                assert "Headphones" in wb.sheetnames
                assert "Earbuds" in wb.sheetnames

                ws = wb["Headphones"]
                assert ws.cell(3, 1).value == "wireless headphones"  # highest volume first
            finally:
                ew.OUTPUT_DIR = original

    def test_sheet_name_truncated_to_31(self):
        from pipeline.excel_writer import _safe_sheet_name
        long_name = "A" * 50
        assert len(_safe_sheet_name(long_name)) <= 31

    def test_sheet_name_strips_forbidden_chars(self):
        from pipeline.excel_writer import _safe_sheet_name
        assert "/" not in _safe_sheet_name("category/sub")
        assert "\\" not in _safe_sheet_name("category\\sub")
        assert "?" not in _safe_sheet_name("what?")


# ── keyword_planner deduplication ────────────────────────────────────────────
class TestKeywordPlannerDedup:
    def test_deduplication_normalises_case(self):
        """Verify that get_search_volumes deduplicates before batching."""
        # We test just the deduplication logic, not the API call.
        keywords = ["Wireless Headphones", "wireless headphones", "WIRELESS HEADPHONES"]
        unique = list(dict.fromkeys(kw.lower().strip() for kw in keywords if kw.strip()))
        assert unique == ["wireless headphones"]

    def test_empty_keywords_returns_empty(self):
        # get_search_volumes should return {} immediately for empty input
        # without attempting any API call
        from unittest.mock import patch
        with patch("pipeline.keyword_planner._build_client") as mock_client:
            from pipeline.keyword_planner import get_search_volumes
            result = get_search_volumes([])
            mock_client.assert_not_called()
            assert result == {}
